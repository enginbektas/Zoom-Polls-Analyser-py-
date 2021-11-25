from pandas.tests.io.excel.test_xlsxwriter import xlsxwriter
from openpyxl import load_workbook
import pyexcel as p

from os import path


def create_statistics(student_list, poll):
    if len(poll.get_questions()) == 0:
        return
    else:
        workbook = xlsxwriter.Workbook('output files/' + poll.get_name() + 'Statistics.xlsx')
        worksheet = workbook.add_worksheet()

        i = 0
        for question in poll.get_questions():
            worksheet.write(i, 0, question.get_text())
            answersDictionary = None
            for student in student_list:
                if len(student.get_answered_polls()) == 0:
                    continue
                saToWork = {}
                for sa in student.get_answered_polls():
                    if sa.get_poll().get_name() == poll.get_name():
                        saToWork = sa

                try:
                    answersDictionary[saToWork.get_answer(question)] += 1
                except:
                    if answersDictionary is None:
                        answersDictionary = dict.fromkeys(range(0), [])
                        answersDictionary[saToWork.get_answer(question)] = 1
                    else:
                        try:
                            answersDictionary[saToWork.get_answer(question)] = 1
                        except KeyError:
                            continue
                    continue
            i += 1
            startIndex = i
            for key in answersDictionary:
                worksheet.write(i, 0, key)
                an = answersDictionary[key]
                worksheet.write(i, 1, an)
                i += 1

            chart1 = workbook.add_chart({'type': 'pie'})
            chart2 = workbook.add_chart({'type': 'column'})
            chart1.add_series({
                'name': 'Pie Chart',
                'categories': ['Sheet1', startIndex, 0, i, 0],
                'values': ['Sheet1', startIndex, 1, i, 1],
            })
            chart2.add_series({
                'name': 'Column Chart',
                'categories': ['Sheet1', startIndex, 0, i, 0],
                'values': ['Sheet1', startIndex, 1, i, 1],
            })
            chart1.set_style(10)
            chart2.set_style(10)
            chart1.set_size({'width': 500, 'height': 250})
            chart2.set_size({'width': 500, 'height': 250})
            worksheet.insert_chart('C' + str(startIndex + 1), chart1, {'x_offset': 0, 'y_offset': 0})
            worksheet.insert_chart('K' + str(startIndex + 1), chart2, {'x_offset': 0, 'y_offset': 0})
            i += 10
        workbook.close()


def create_attendance_output(student_list):
    if path.exists('output files/CSE3063_Fall2020_Attendance_List.xlsx'):
        pass
    else:
        p.save_book_as(file_name='excel files/CES3063_Fall2020_rptSinifListesi.xls',
                       dest_file_name='output files/CSE3063_Fall2020_Attendance_List.xlsx')

    wb = load_workbook('output files/CSE3063_Fall2020_Attendance_List.xlsx')
    ws = wb.worksheets[0]
    i = 14
    ws['L13'] = "Attendance polls"
    ws['M13'] = "Attendance rate"
    ws['N13'] = "Attendance percentage"
    for student in student_list:
        ws['L' + str(i)] = str(student.get_attendance())
        ws['M' + str(i)] = str(student.get_attendance()) + " of " + str(student.get_totalAttendance())
        if student.get_totalAttendance() != 0:
            ws['N' + str(i)] = str(student.get_attendance() / student.get_totalAttendance() * 100)
        else:
            ws['N' + str(i)] = "0"

        if i == 204 or i == 209:
            i += 5
        elif i == 215:
            i += 15
        else:
            i += 1
    wb.save('output files/CSE3063_Fall2020_Attendance_List.xlsx')


def create_poll_output(student_list, poll):
    if len(poll.get_questions()) == 0:
        return
    else:
        if path.exists('output files/' + poll.get_name() + '.xlsx'):
            pass
        else:
            p.save_book_as(file_name='excel files/CES3063_Fall2020_rptSinifListesi.xls',
                           dest_file_name='output files/' + poll.get_name() + '.xlsx')

        wb = load_workbook('output files/' + poll.get_name() + '.xlsx')
        ws = wb.worksheets[0]
        len_poll = len(poll.get_questions())

        i = 14
        column_chr = 76
        for j in range(1, len_poll + 1):
            ws[chr(column_chr) + '13'] = "Q" + str(j)
            j += 1
            column_chr += 1
        success_chr = column_chr
        column_chr = 76
        ws[chr(success_chr) + '13'] = "Number of questions"
        ws[chr(success_chr + 1) + '13'] = "Success Percentage"
        count2 = 0
        for student in student_list:
            count2 += 1
            answered_poll = None
            count = 0
            correct_answer = 0
            if student.get_answered_polls() is None or len(student.get_answered_polls()) == 0:
                while success_chr > column_chr:
                    ws[chr(column_chr) + str(i)] = "0"
                    column_chr += 1
                    count += 1
            else:
                for ap in student.get_answered_polls():
                    if ap.get_poll() is None:
                        break
                    if poll.get_name() == ap.get_poll().get_name():
                        answered_poll = ap
                        break

                while success_chr > column_chr:
                    try:
                        if answered_poll is None:
                            break

                        if answered_poll.get_answer(poll.get_questions()[count]) == poll.get_questions()[
                            count].get_trueChoice():

                            ws[chr(column_chr) + str(i)] = "1"
                            correct_answer += 1
                        else:
                            ws[chr(column_chr) + str(i)] = "0"
                        column_chr += 1
                        count += 1
                    except KeyError:
                        break
            column_chr = 76
            ws[chr(success_chr) + str(i)] = str(success_chr - column_chr)
            ws[chr(success_chr + 1) + str(i)] = str(correct_answer / (success_chr - column_chr) * 100)
            if i == 209:
                i += 5
            elif i == 204:
                i += 5
            elif i == 215:
                i += 15
            else:
                i += 1

        wb.save('output files/' + poll.get_name() + '.xlsx')


def create_global_output(student_list, poll_list):
    if path.exists('output files/CSE33063_Fall2020_Global_Output.xlsx'):
        pass
    else:
        p.save_book_as(file_name='excel files/CES3063_Fall2020_rptSinifListesi.xls',
                       dest_file_name='output files/CSE33063_Fall2020_Global_Output.xlsx')
    wb = load_workbook('output files/CSE33063_Fall2020_Global_Output.xlsx')
    ws = wb.worksheets[0]
    column_chr = 78

    j = 0
    for poll in poll_list:
        first_column = chr(column_chr)
        second_column = chr(column_chr + 1)
        third_column = chr(column_chr + 2)
        ws[first_column + '13'] = "q" + str(j) + "_name"
        ws[second_column + '13'] = "q" + str(j) + "_numOfQuestions"
        ws[third_column + '13'] = "q" + str(j) + "successRate"

        i = 14
        for student in student_list:
            ws[first_column + str(i)] = str(poll.get_name())
            ws[second_column + str(i)] = str(len(poll.get_questions()))

            answered_poll = None
            correct_answer = 0
            num_of_questions = 0
            for ap in student.get_answered_polls():  # to find to student answer
                if poll.get_name() == ap.get_poll().get_name():
                    answered_poll = ap
                    break
            for q in poll.get_questions():  # find to number of questions and correct answer
                if q is None or answered_poll is None or len(answered_poll.get_questions_and_answers()) == 0:
                    break
                num_of_questions += 1
                if answered_poll.get_answer(q) == q.get_trueChoice():
                    correct_answer += 1
            if num_of_questions == 0:
                ws[third_column + str(i)] = "0"
            else:
                ws[third_column + str(i)] = str(correct_answer / num_of_questions * 100)
            if i == 209:
                i += 5
            elif i == 204:
                i += 5
            elif i == 215:
                i += 15
            else:
                i += 1
        column_chr += 3
    wb.save('output files/CSE33063_Fall2020_Global_Output.xlsx')
